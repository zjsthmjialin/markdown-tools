from .base import BaseConverter
from openpyxl import load_workbook

class XlsxConverter(BaseConverter):
    def get_supported_extensions(self) -> list[str]:
        return ['.xlsx']

    def to_markdown(self, file_path: str, output_dir: str) -> str:
        wb = load_workbook(file_path, data_only=True)
        markdown_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            markdown_parts.append(f"## {sheet_name}\n")

            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else '' for cell in row]
                rows.append('| ' + ' | '.join(cells) + ' |')

            if rows:
                # 添加表头分隔符
                col_count = len(rows[0].split('|')[:-1]) if rows else 0
                rows.insert(1, '| ' + ' | '.join(['---'] * col_count) + ' |')

            markdown_parts.append('\n'.join(rows))
            markdown_parts.append('')

        return '\n\n'.join(markdown_parts)